import time
import datetime
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys


s = Service("E:\\РАБОЧИЕ ДОКУМЕНТЫ\\документы\\иницативы техсовет\\2021\\конструктор\\2 TS s TE\\chromedriver.exe")

path = "123.xlsx"  # имя файла
n = input(str("Введите номер протокола: "))
date_prot = input('введите дату проведения протокола (гггг-мм-дд): ')
path_protokol = input('введите путь к протоколу: ')
date_now = datetime.date.today()

date_prot = date_prot.split('-')
day_count = datetime.date(int(date_prot[0]), int(date_prot[1]), int(date_prot[2]))
day_count = date_now - day_count
day_count = str(day_count)
day_count = (day_count.split()[0])

wb_obj = openpyxl.load_workbook(path)  # Открываем файл
sheet_obj = wb_obj.active  # Выбираем активный лист таблицы(
m_row = sheet_obj.max_row

url = "https://ideas.nlmk.com/"

options = webdriver.ChromeOptions()
options.add_argument("user-agent=Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:84.0) Gecko/20100101 Firefox/84.0")
options.headless = False

driver = webdriver.Chrome(
    service=s,
    options=options
)

try:
    driver.get(url=url)
    time.sleep(15)
    '''
    email_input = driver.find_element(By.ID, "login")
    email_input.clear()
    email_input.send_keys("odinokov_da")

    password_input = driver.find_element(By.ID, "password")
    password_input.clear()
    password_input.send_keys(id_password)
    time.sleep(3)
    password_input.send_keys(Keys.ENTER)
    time.sleep(6)
    '''
    print('вход')
    driver.find_element(By.CLASS_NAME, "work").click()
    time.sleep(5)
    print('раб')
    driver.find_element(By.LINK_TEXT, "Все идеи").click()
    time.sleep(5)
    print('все')
    driver.find_element(By.LINK_TEXT, "2ТС ТЭ").click()
    time.sleep(5)
    print('2ТС ТЭ')
    """
    оперделяем есть ли страницы если есть,определяем количество страниц
    и заносим их в список 
    """
    my_ul2 = driver.find_elements(By.XPATH, "//ul[@class='b-page']")  # блок со страницам
    if len(my_ul2)==1:
        my_ul = driver.find_element(By.XPATH, "//ul[@class='b-page']")  # блок со страницам
        time.sleep(7)
        all_li = my_ul.find_elements(By.TAG_NAME, "li")  # оперделяем кол-во страниц
        time.sleep(7)
        link_list = []
        for li in all_li:
            y = li.text
            link_list.append(y)

        rec = 1
        while rec == 1:
            for i in range(2, m_row + 1):
                cell_obj = sheet_obj.cell(row=i, column=1)  # В column= подставляем номер нужной колонки
                number_ideas = cell_obj.value
                cell_ob = sheet_obj.cell(row=i, column=10)  # В column= подставляем номер нужной колонки
                name_ideas = cell_ob.value

                for y in link_list:
                    y = y
                    ds = driver.find_elements(By.XPATH, " // *[ @ href = 'Offer.aspx?id=" + str(number_ideas) + "']")
                    time.sleep(7)

                    if len(ds) >= 1:
                        driver.find_element(By.XPATH, " // *[ @ href = 'Offer.aspx?id=" + str(number_ideas) + "']").click()
                        time.sleep(8)
                        """
                        заходим во вкладку 2ТС
                        """

                        driver.find_element(By.ID,
                                            "ctl00_ctl00_ctl00_MainContent_VMenuRightContent_OfferMenu_linkEngine83") \
                            .click()
                        time.sleep(8)
                        driver.find_element(By.ID,
                                            "ctl00_ctl00_ctl00_MainContent_CommandsMenu_linkEdit").click()
                        time.sleep(5)
                        '''
                        делее место заполнения № протокола 2ТС
                        и даты принятия протокола находятся в 
                        iframe, заходим в него и вносим данные
                        '''
                        driver.switch_to.frame(driver.find_element(By.TAG_NAME, "iframe"))
                        time.sleep(5)
                        input_n_prot = driver.find_element(By.ID, "alpaca2")
                        time.sleep(5)
                        input_n_prot.clear()
                        input_n_prot.send_keys(n)
                        time.sleep(5)

                        '''
                        вносим дату в поле методом перемещения курсора стрелкам
                        '''

                        driver.find_element(By.XPATH,
                                            "// input[@class='alpaca-control form-control' and @id = 'alpaca3']").clear()
                        driver.find_element(By.XPATH,
                                            "// input[@class='alpaca-control form-control' and @id = 'alpaca3']").send_keys(
                            Keys.LEFT * int(day_count))
                        driver.find_element(By.XPATH,
                                            "// input[@class='alpaca-control form-control' and @id = 'alpaca4']").clear()
                        driver.find_element(By.XPATH,
                                            "// input[@class='alpaca-control form-control' and @id = 'alpaca4']").send_keys(
                            Keys.LEFT * int(day_count))
                        #  time.sleep(5)

                        driver.find_element(By.XPATH,
                                            "//select[contains(@id,'alpaca5')]").send_keys(Keys.DOWN)
                        #  time.sleep(5)

                        driver.find_element(By.XPATH,
                                            "//select[contains(@id,'alpaca12')]").send_keys(Keys.DOWN)
                        time.sleep(3)

                        '''
                        выходим из iframe для сохранения заполненных
                         результатов по средсвом witch_to.default_content()
                        '''

                        driver.switch_to.default_content()
                        time.sleep(5)
                        driver.find_element(By.ID,
                                            "ctl00_ctl00_ctl00_MainContent_CommandsMenu_linkSave").click()
                        time.sleep(5)
                        driver.find_element(By.ID,
                                            "ctl00_ctl00_ctl00_MainContent_VMenuRightContent_"
                                            "OfferMenu_linkDocuments").click()

                        """
                        передаем файл во вкладку документы
                        """

                        time.sleep(5)
                        driver.find_element(By.XPATH, "//input[@type='file'][contains(@id,"
                                                      "'FileUpload1')]").send_keys(path_protokol)
                        time.sleep(5)
                        driver.find_element(By.ID,
                                            "ctl00_ctl00_ctl00_MainContent_VMenuLeftContent"
                                            "_OfferContent_AttachedFilesCore_BtnSave").click()
                        time.sleep(5)

                        '''
                                 заходим в журнал статусов
                                '''

                        
                        driver.find_element(By.ID,
                                            "ctl00_ctl00_ctl00_MainContent_VMenuRightContent_OfferMenu_linkStatus").click()
                        time.sleep(5)

                        driver.find_element(By.ID, "ctl00_ctl00_ctl00_MainContent_CommandsMenu_linkEdit").click()
                        time.sleep(10)
                        driver.find_element(By.XPATH,
                                            "// *[@class='ddlStatusList' and @id = "
                                            "'ctl00_ctl00_ctl00_MainContent_VMenuLeftContent_"
                                            "OfferContent_newStatusList']") \
                            .send_keys(Keys.DOWN)
                        time.sleep(10)
                        driver.find_element(By.ID,
                                            "ctl00_ctl00_ctl00_MainContent_VMenuLeftContent"
                                            "_OfferContent_changeStatusButton") \
                            .click()
                        time.sleep(10)
                        
                        driver.find_element(By.LINK_TEXT, "Все идеи").click()
                        time.sleep(10)

                        print(name_ideas)
                        break
                    else:
                        rec = 0
                        w = int(y) + 1
                        print(f"переходим на страницу: {w}")
                        driver.find_element(By.XPATH, " // *[ @ href = 'SearchOffers.aspx?page=" + str(y) + "']").click()
                        time.sleep(10)
    
    else:
        for i in range(2, m_row + 1):
                cell_obj = sheet_obj.cell(row=i, column=1)  # В column= подставляем номер нужной колонки
                number_ideas = cell_obj.value
                cell_ob = sheet_obj.cell(row=i, column=10)  # В column= подставляем номер нужной колонки
                name_ideas = cell_ob.value
                driver.find_element(By.XPATH, " // *[ @ href = 'Offer.aspx?id=" + str(number_ideas) + "']").click()
                time.sleep(8)
                """
                заходим во вкладку 2ТС
                """

                driver.find_element(By.ID,
                                    "ctl00_ctl00_ctl00_MainContent_VMenuRightContent_OfferMenu_linkEngine83") \
                    .click()
                time.sleep(8)
                driver.find_element(By.ID,
                                    "ctl00_ctl00_ctl00_MainContent_CommandsMenu_linkEdit").click()
                time.sleep(5)
                '''
                делее место заполнения № протокола 2ТС
                и даты принятия протокола находятся в 
                iframe, заходим в него и вносим данные
                '''
                driver.switch_to.frame(driver.find_element(By.TAG_NAME, "iframe"))
                time.sleep(5)
                input_n_prot = driver.find_element(By.ID, "alpaca2")
                time.sleep(5)
                input_n_prot.clear()
                input_n_prot.send_keys(n)
                time.sleep(5)

                '''
                вносим дату в поле методом перемещения курсора стрелкам
                '''

                driver.find_element(By.XPATH,
                                    "// input[@class='alpaca-control form-control' and @id = 'alpaca3']").clear()
                driver.find_element(By.XPATH,
                                    "// input[@class='alpaca-control form-control' and @id = 'alpaca3']").send_keys(
                    Keys.LEFT * int(day_count))
                driver.find_element(By.XPATH,
                                    "// input[@class='alpaca-control form-control' and @id = 'alpaca4']").clear()
                driver.find_element(By.XPATH,
                                    "// input[@class='alpaca-control form-control' and @id = 'alpaca4']").send_keys(
                    Keys.LEFT * int(day_count))
                #  time.sleep(5)

                driver.find_element(By.XPATH,
                                    "//select[contains(@id,'alpaca5')]").send_keys(Keys.DOWN)
                #  time.sleep(5)

                driver.find_element(By.XPATH,
                                    "//select[contains(@id,'alpaca12')]").send_keys(Keys.DOWN)
                time.sleep(3)

                '''
                выходим из iframe для сохранения заполненных
                результатов по средсвом witch_to.default_content()
                '''

                driver.switch_to.default_content()
                time.sleep(5)
                driver.find_element(By.ID,
                                    "ctl00_ctl00_ctl00_MainContent_CommandsMenu_linkSave").click()
                time.sleep(5)
                driver.find_element(By.ID,
                                    "ctl00_ctl00_ctl00_MainContent_VMenuRightContent_"
                                    "OfferMenu_linkDocuments").click()

                """
                передаем файл во вкладку документы
                """

                time.sleep(5)
                driver.find_element(By.XPATH, "//input[@type='file'][contains(@id,"
                                              "'FileUpload1')]").send_keys(path_protokol)
                time.sleep(5)
                driver.find_element(By.ID,
                                    "ctl00_ctl00_ctl00_MainContent_VMenuLeftContent"
                                    "_OfferContent_AttachedFilesCore_BtnSave").click()
                time.sleep(5)

                '''
                заходим в журнал статусов
                '''

                
                driver.find_element(By.ID,
                                    "ctl00_ctl00_ctl00_MainContent_VMenuRightContent_OfferMenu_linkStatus").click()
                time.sleep(5)

                driver.find_element(By.ID, "ctl00_ctl00_ctl00_MainContent_CommandsMenu_linkEdit").click()
                time.sleep(10)
                driver.find_element(By.XPATH,
                                    "// *[@class='ddlStatusList' and @id = "
                                    "'ctl00_ctl00_ctl00_MainContent_VMenuLeftContent_"
                                    "OfferContent_newStatusList']") \
                    .send_keys(Keys.DOWN)
                time.sleep(10)
                driver.find_element(By.ID,
                                    "ctl00_ctl00_ctl00_MainContent_VMenuLeftContent"
                                    "_OfferContent_changeStatusButton") \
                    .click()
                time.sleep(10)
                
                driver.find_element(By.LINK_TEXT, "Все идеи").click()
                time.sleep(10)

                print(name_ideas)
except Exception as ex:
    print(ex)
finally:
    driver.close()
    driver.quit()
